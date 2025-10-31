"""
Generador de Excel para Reportes
Utiliza openpyxl para crear documentos Excel
"""

from io import BytesIO
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  WARNING: openpyxl no está instalado. Instala con: pip install openpyxl")


class ExcelGenerator:
    """Generador de reportes en Excel"""
    
    # Estilos predefinidos
    HEADER_FILL = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT = Font(bold=True, size=14)
    SUBTITLE_FONT = Font(size=10, italic=True)
    
    @staticmethod
    def _check_openpyxl():
        """Verifica si openpyxl está disponible"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl no está instalado. Ejecuta: pip install openpyxl")
    
    @staticmethod
    def _apply_header_style(ws, row, columns):
        """Aplica estilo a la fila de encabezado"""
        for col_num, _ in enumerate(columns, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.fill = ExcelGenerator.HEADER_FILL
            cell.font = ExcelGenerator.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    @staticmethod
    def _auto_adjust_columns(ws):
        """Ajusta automáticamente el ancho de las columnas"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def generar_reporte_ventas(datos, fecha_inicio=None, fecha_fin=None):
        """
        Genera Excel de reporte de ventas
        
        Args:
            datos (dict): Datos del reporte
            fecha_inicio (str): Fecha inicio
            fecha_fin (str): Fecha fin
            
        Returns:
            BytesIO: Buffer con el Excel generado
        """
        ExcelGenerator._check_openpyxl()
        
        wb = Workbook()
        
        # HOJA 1: RESUMEN GENERAL
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        
        # Título
        ws_resumen['A1'] = "REPORTE DE VENTAS - REVENGE POS"
        ws_resumen['A1'].font = ExcelGenerator.TITLE_FONT
        
        # Período
        periodo = f"Del {fecha_inicio} al {fecha_fin}" if fecha_inicio and fecha_fin else "Todas las ventas"
        ws_resumen['A2'] = f"Período: {periodo}"
        ws_resumen['A2'].font = ExcelGenerator.SUBTITLE_FONT
        
        ws_resumen['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws_resumen['A3'].font = ExcelGenerator.SUBTITLE_FONT
        
        # Resumen
        resumen = datos.get('resumen', {})
        ws_resumen['A5'] = "Indicador"
        ws_resumen['B5'] = "Valor"
        ExcelGenerator._apply_header_style(ws_resumen, 5, ['A', 'B'])
        
        ws_resumen['A6'] = "Total de Ventas (S/)"
        ws_resumen['B6'] = resumen.get('monto_total', 0)
        ws_resumen['A7'] = "Número de Transacciones"
        ws_resumen['B7'] = resumen.get('total_ventas', 0)
        ws_resumen['A8'] = "Ticket Promedio"
        ws_resumen['B8'] = resumen.get('promedio_venta', 0)
        ws_resumen['A9'] = "Venta Mínima"
        ws_resumen['B9'] = resumen.get('venta_minima', 0)
        ws_resumen['A10'] = "Venta Máxima"
        ws_resumen['B10'] = resumen.get('venta_maxima', 0)
        
        # Formato de moneda
        for row in range(6, 11):
            ws_resumen[f'B{row}'].number_format = 'S/ #,##0.00'
        
        ExcelGenerator._auto_adjust_columns(ws_resumen)
        
        # HOJA 2: PRODUCTOS MÁS VENDIDOS
        if 'productos_mas_vendidos' in datos and datos['productos_mas_vendidos']:
            ws_productos = wb.create_sheet("Productos")
            
            ws_productos['A1'] = "DETALLE DE VENTAS POR PRODUCTO"
            ws_productos['A1'].font = ExcelGenerator.TITLE_FONT
            
            headers = ['Código', 'Producto', 'Categoría', 'Cantidad', 'Precio Unit.', 'Total', '% Part.']
            for col_num, header in enumerate(headers, 1):
                ws_productos.cell(row=3, column=col_num, value=header)
            ExcelGenerator._apply_header_style(ws_productos, 3, headers)
            
            for row_num, producto in enumerate(datos['productos_mas_vendidos'], 4):
                ws_productos.cell(row=row_num, column=1, value=producto.get('codigo', ''))
                ws_productos.cell(row=row_num, column=2, value=producto.get('producto', ''))
                ws_productos.cell(row=row_num, column=3, value=producto.get('categoria', ''))
                ws_productos.cell(row=row_num, column=4, value=producto.get('cantidad', 0))
                ws_productos.cell(row=row_num, column=5, value=producto.get('precio_unitario', 0))
                ws_productos.cell(row=row_num, column=6, value=producto.get('total', 0))
                ws_productos.cell(row=row_num, column=7, value=producto.get('porcentaje', 0))
                
                ws_productos.cell(row=row_num, column=5).number_format = 'S/ #,##0.00'
                ws_productos.cell(row=row_num, column=6).number_format = 'S/ #,##0.00'
                ws_productos.cell(row=row_num, column=7).number_format = '0.00"%"'
            
            ExcelGenerator._auto_adjust_columns(ws_productos)
        
        # HOJA 3: VENTAS POR CAJERO
        if 'ventas_por_cajero' in datos and datos['ventas_por_cajero']:
            ws_cajeros = wb.create_sheet("Cajeros")
            
            ws_cajeros['A1'] = "VENTAS POR CAJERO"
            ws_cajeros['A1'].font = ExcelGenerator.TITLE_FONT
            
            headers = ['Cajero', 'Ventas Totales', 'Nº Trans.', 'Ticket Prom.', '% del Total']
            for col_num, header in enumerate(headers, 1):
                ws_cajeros.cell(row=3, column=col_num, value=header)
            ExcelGenerator._apply_header_style(ws_cajeros, 3, headers)
            
            for row_num, cajero in enumerate(datos['ventas_por_cajero'], 4):
                ws_cajeros.cell(row=row_num, column=1, value=cajero.get('cajero', ''))
                ws_cajeros.cell(row=row_num, column=2, value=cajero.get('total', 0))
                ws_cajeros.cell(row=row_num, column=3, value=cajero.get('cantidad', 0))
                ws_cajeros.cell(row=row_num, column=4, value=cajero.get('ticket_promedio', 0))
                ws_cajeros.cell(row=row_num, column=5, value=cajero.get('porcentaje', 0))
                
                ws_cajeros.cell(row=row_num, column=2).number_format = 'S/ #,##0.00'
                ws_cajeros.cell(row=row_num, column=4).number_format = 'S/ #,##0.00'
                ws_cajeros.cell(row=row_num, column=5).number_format = '0.00"%"'
            
            ExcelGenerator._auto_adjust_columns(ws_cajeros)
        
        # HOJA 4: VENTAS POR DÍA
        if 'ventas_por_dia' in datos and datos['ventas_por_dia']:
            ws_dias = wb.create_sheet("Por Día")
            
            ws_dias['A1'] = "VENTAS POR DÍA"
            ws_dias['A1'].font = ExcelGenerator.TITLE_FONT
            
            headers = ['Fecha', 'Ventas Totales', 'Nº Transacciones']
            for col_num, header in enumerate(headers, 1):
                ws_dias.cell(row=3, column=col_num, value=header)
            ExcelGenerator._apply_header_style(ws_dias, 3, headers)
            
            for row_num, dia in enumerate(datos['ventas_por_dia'], 4):
                ws_dias.cell(row=row_num, column=1, value=dia.get('fecha', ''))
                ws_dias.cell(row=row_num, column=2, value=dia.get('total', 0))
                ws_dias.cell(row=row_num, column=3, value=dia.get('cantidad', 0))
                
                ws_dias.cell(row=row_num, column=2).number_format = 'S/ #,##0.00'
            
            ExcelGenerator._auto_adjust_columns(ws_dias)
        
        # HOJA 5: MÉTODOS DE PAGO
        if 'ventas_por_metodo_pago' in datos and datos['ventas_por_metodo_pago']:
            ws_metodos = wb.create_sheet("Métodos de Pago")
            
            ws_metodos['A1'] = "VENTAS POR MÉTODO DE PAGO"
            ws_metodos['A1'].font = ExcelGenerator.TITLE_FONT
            
            headers = ['Método', 'Ventas Totales', '%']
            for col_num, header in enumerate(headers, 1):
                ws_metodos.cell(row=3, column=col_num, value=header)
            ExcelGenerator._apply_header_style(ws_metodos, 3, headers)
            
            for row_num, metodo in enumerate(datos['ventas_por_metodo_pago'], 4):
                ws_metodos.cell(row=row_num, column=1, value=metodo.get('metodo', ''))
                ws_metodos.cell(row=row_num, column=2, value=metodo.get('total', 0))
                ws_metodos.cell(row=row_num, column=3, value=metodo.get('porcentaje', 0))
                
                ws_metodos.cell(row=row_num, column=2).number_format = 'S/ #,##0.00'
                ws_metodos.cell(row=row_num, column=3).number_format = '0.00"%"'
            
            ExcelGenerator._auto_adjust_columns(ws_metodos)
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def generar_reporte_compras(datos, fecha_inicio=None, fecha_fin=None):
        """Genera Excel de reporte de compras"""
        ExcelGenerator._check_openpyxl()
        
        wb = Workbook()
        
        # HOJA 1: RESUMEN
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        
        ws_resumen['A1'] = "REPORTE DE COMPRAS - REVENGE POS"
        ws_resumen['A1'].font = ExcelGenerator.TITLE_FONT
        
        periodo = f"Del {fecha_inicio} al {fecha_fin}" if fecha_inicio and fecha_fin else "Todas las compras"
        ws_resumen['A2'] = f"Período: {periodo}"
        ws_resumen['A2'].font = ExcelGenerator.SUBTITLE_FONT
        
        ws_resumen['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws_resumen['A3'].font = ExcelGenerator.SUBTITLE_FONT
        
        resumen = datos.get('resumen', {})
        ws_resumen['A5'] = "Indicador"
        ws_resumen['B5'] = "Valor"
        ExcelGenerator._apply_header_style(ws_resumen, 5, ['A', 'B'])
        
        ws_resumen['A6'] = "Total de Compras (S/)"
        ws_resumen['B6'] = resumen.get('monto_total', 0)
        ws_resumen['A7'] = "Número de Transacciones"
        ws_resumen['B7'] = resumen.get('total_compras', 0)
        ws_resumen['A8'] = "Promedio por Compra"
        ws_resumen['B8'] = resumen.get('promedio_compra', 0)
        
        for row in range(6, 9):
            ws_resumen[f'B{row}'].number_format = 'S/ #,##0.00'
        
        ExcelGenerator._auto_adjust_columns(ws_resumen)
        
        # HOJA 2: PRODUCTOS
        if 'productos_mas_comprados' in datos and datos['productos_mas_comprados']:
            ws_productos = wb.create_sheet("Productos")
            
            ws_productos['A1'] = "DETALLE DE COMPRAS POR PRODUCTO"
            ws_productos['A1'].font = ExcelGenerator.TITLE_FONT
            
            headers = ['Código', 'Producto', 'Categoría', 'Cantidad', 'Precio Unit.', 'Total', '% Part.']
            for col_num, header in enumerate(headers, 1):
                ws_productos.cell(row=3, column=col_num, value=header)
            ExcelGenerator._apply_header_style(ws_productos, 3, headers)
            
            for row_num, producto in enumerate(datos['productos_mas_comprados'], 4):
                ws_productos.cell(row=row_num, column=1, value=producto.get('codigo', ''))
                ws_productos.cell(row=row_num, column=2, value=producto.get('producto', ''))
                ws_productos.cell(row=row_num, column=3, value=producto.get('categoria', ''))
                ws_productos.cell(row=row_num, column=4, value=producto.get('cantidad', 0))
                ws_productos.cell(row=row_num, column=5, value=producto.get('precio_unitario', 0))
                ws_productos.cell(row=row_num, column=6, value=producto.get('total', 0))
                ws_productos.cell(row=row_num, column=7, value=producto.get('porcentaje', 0))
                
                ws_productos.cell(row=row_num, column=5).number_format = 'S/ #,##0.00'
                ws_productos.cell(row=row_num, column=6).number_format = 'S/ #,##0.00'
                ws_productos.cell(row=row_num, column=7).number_format = '0.00"%"'
            
            ExcelGenerator._auto_adjust_columns(ws_productos)
        
        # HOJA 3: PROVEEDORES
        if 'compras_por_proveedor' in datos and datos['compras_por_proveedor']:
            ws_proveedores = wb.create_sheet("Proveedores")
            
            ws_proveedores['A1'] = "COMPRAS POR PROVEEDOR"
            ws_proveedores['A1'].font = ExcelGenerator.TITLE_FONT
            
            headers = ['Proveedor', 'Nº Compras', 'Total', '% del Total']
            for col_num, header in enumerate(headers, 1):
                ws_proveedores.cell(row=3, column=col_num, value=header)
            ExcelGenerator._apply_header_style(ws_proveedores, 3, headers)
            
            for row_num, proveedor in enumerate(datos['compras_por_proveedor'], 4):
                ws_proveedores.cell(row=row_num, column=1, value=proveedor.get('proveedor', ''))
                ws_proveedores.cell(row=row_num, column=2, value=proveedor.get('cantidad', 0))
                ws_proveedores.cell(row=row_num, column=3, value=proveedor.get('total', 0))
                ws_proveedores.cell(row=row_num, column=4, value=proveedor.get('porcentaje', 0))
                
                ws_proveedores.cell(row=row_num, column=3).number_format = 'S/ #,##0.00'
                ws_proveedores.cell(row=row_num, column=4).number_format = '0.00"%"'
            
            ExcelGenerator._auto_adjust_columns(ws_proveedores)
        
        # HOJA 4: POR DÍA
        if 'compras_por_dia' in datos and datos['compras_por_dia']:
            ws_dias = wb.create_sheet("Por Día")
            
            ws_dias['A1'] = "COMPRAS POR DÍA"
            ws_dias['A1'].font = ExcelGenerator.TITLE_FONT
            
            headers = ['Fecha', 'Total Compras', 'Nº Transacciones']
            for col_num, header in enumerate(headers, 1):
                ws_dias.cell(row=3, column=col_num, value=header)
            ExcelGenerator._apply_header_style(ws_dias, 3, headers)
            
            for row_num, dia in enumerate(datos['compras_por_dia'], 4):
                ws_dias.cell(row=row_num, column=1, value=dia.get('fecha', ''))
                ws_dias.cell(row=row_num, column=2, value=dia.get('total', 0))
                ws_dias.cell(row=row_num, column=3, value=dia.get('cantidad', 0))
                
                ws_dias.cell(row=row_num, column=2).number_format = 'S/ #,##0.00'
            
            ExcelGenerator._auto_adjust_columns(ws_dias)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def generar_reporte_inventario(datos):
        """Genera Excel de reporte de inventario"""
        ExcelGenerator._check_openpyxl()
        
        wb = Workbook()
        
        # HOJA 1: RESUMEN
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        
        ws_resumen['A1'] = "REPORTE DE INVENTARIO - REVENGE POS"
        ws_resumen['A1'].font = ExcelGenerator.TITLE_FONT
        
        ws_resumen['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws_resumen['A2'].font = ExcelGenerator.SUBTITLE_FONT
        
        resumen = datos.get('resumen', {})
        ws_resumen['A4'] = "Indicador"
        ws_resumen['B4'] = "Valor"
        ExcelGenerator._apply_header_style(ws_resumen, 4, ['A', 'B'])
        
        ws_resumen['A5'] = "Total de Productos"
        ws_resumen['B5'] = resumen.get('total_productos', 0)
        ws_resumen['A6'] = "Valor del Inventario"
        ws_resumen['B6'] = resumen.get('valor_inventario', 0)
        ws_resumen['A7'] = "Valor Potencial de Venta"
        ws_resumen['B7'] = resumen.get('valor_venta_potencial', 0)
        ws_resumen['A8'] = "Productos con Stock Bajo"
        ws_resumen['B8'] = resumen.get('productos_stock_bajo', 0)
        ws_resumen['A9'] = "Productos Sin Stock"
        ws_resumen['B9'] = resumen.get('productos_sin_stock', 0)
        
        for row in range(6, 8):
            ws_resumen[f'B{row}'].number_format = 'S/ #,##0.00'
        
        ExcelGenerator._auto_adjust_columns(ws_resumen)
        
        # HOJA 2: POR CATEGORÍA
        if 'productos_por_categoria' in datos and datos['productos_por_categoria']:
            ws_categorias = wb.create_sheet("Por Categoría")
            
            ws_categorias['A1'] = "INVENTARIO POR CATEGORÍA"
            ws_categorias['A1'].font = ExcelGenerator.TITLE_FONT
            
            headers = ['Categoría', 'Cant. Productos', 'Stock Total', 'Valor']
            for col_num, header in enumerate(headers, 1):
                ws_categorias.cell(row=3, column=col_num, value=header)
            ExcelGenerator._apply_header_style(ws_categorias, 3, headers)
            
            for row_num, categoria in enumerate(datos['productos_por_categoria'], 4):
                ws_categorias.cell(row=row_num, column=1, value=categoria.get('categoria', ''))
                ws_categorias.cell(row=row_num, column=2, value=categoria.get('cantidad', 0))
                ws_categorias.cell(row=row_num, column=3, value=categoria.get('stock_total', 0))
                ws_categorias.cell(row=row_num, column=4, value=categoria.get('valor', 0))
                
                ws_categorias.cell(row=row_num, column=4).number_format = 'S/ #,##0.00'
            
            ExcelGenerator._auto_adjust_columns(ws_categorias)
        
        # HOJA 3: STOCK BAJO
        if 'productos_stock_bajo' in datos and datos['productos_stock_bajo']:
            ws_stock_bajo = wb.create_sheet("Stock Bajo")
            
            ws_stock_bajo['A1'] = "PRODUCTOS CON STOCK BAJO"
            ws_stock_bajo['A1'].font = ExcelGenerator.TITLE_FONT
            
            headers = ['Código', 'Producto', 'Categoría', 'Stock Actual', 'Stock Mínimo']
            for col_num, header in enumerate(headers, 1):
                ws_stock_bajo.cell(row=3, column=col_num, value=header)
            ExcelGenerator._apply_header_style(ws_stock_bajo, 3, headers)
            
            for row_num, producto in enumerate(datos['productos_stock_bajo'], 4):
                ws_stock_bajo.cell(row=row_num, column=1, value=producto.get('codigo_barras', ''))
                ws_stock_bajo.cell(row=row_num, column=2, value=producto.get('nombre', ''))
                ws_stock_bajo.cell(row=row_num, column=3, value=producto.get('categoria', ''))
                ws_stock_bajo.cell(row=row_num, column=4, value=producto.get('stock_actual', 0))
                ws_stock_bajo.cell(row=row_num, column=5, value=producto.get('stock_minimo', 0))
            
            ExcelGenerator._auto_adjust_columns(ws_stock_bajo)
        
        # HOJA 4: SIN STOCK
        if 'productos_sin_stock' in datos and datos['productos_sin_stock']:
            ws_sin_stock = wb.create_sheet("Sin Stock")
            
            ws_sin_stock['A1'] = "PRODUCTOS SIN STOCK"
            ws_sin_stock['A1'].font = ExcelGenerator.TITLE_FONT
            
            headers = ['Código', 'Producto', 'Categoría']
            for col_num, header in enumerate(headers, 1):
                ws_sin_stock.cell(row=3, column=col_num, value=header)
            ExcelGenerator._apply_header_style(ws_sin_stock, 3, headers)
            
            for row_num, producto in enumerate(datos['productos_sin_stock'], 4):
                ws_sin_stock.cell(row=row_num, column=1, value=producto.get('codigo_barras', ''))
                ws_sin_stock.cell(row=row_num, column=2, value=producto.get('nombre', ''))
                ws_sin_stock.cell(row=row_num, column=3, value=producto.get('categoria', ''))
            
            ExcelGenerator._auto_adjust_columns(ws_sin_stock)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
